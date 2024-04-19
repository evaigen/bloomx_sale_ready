import requests
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles.colors import WHITE, RGB
import warnings

# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# FIXING THE ARGB HEX VALUES ERROR
__old_rgb_set__ = RGB.__set__

# Dictionary of markings
bloomx_codes = {
    'ufa': ['Зикрач Армавир', 0, 4.0, 1.02],
    'marf': ['Геворгян', 0, 4.0, 1.02]
}

# XPath expression to extract the currency rate
x_euro = '//*[@id="content"]/div/div/div/div[3]/div/table/tbody/tr[16]/td[5]'


def __rgb_set_fixed__(self, instance, value):
    try:
        __old_rgb_set__(self, instance, value)
    except ValueError as e:
        if e.args[0] == 'Colors must be aRGB hex values':
            __old_rgb_set__(self, instance, WHITE)


def parsing_currency():
    # URL of the currency website
    url = "https://www.cbr.ru/currency_base/daily/"

    # Send a GET request to the website
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        html_content = response.text
        parser = etree.HTMLParser()
        tree = etree.fromstring(html_content, parser)

        # Find the currency rate element using XPath
        euro_element = tree.xpath(x_euro)

        # Check if the element was found
        if euro_element:
            # Get the text content of the element
            euro_rate = euro_element[0].text
            euro_rate = euro_rate.replace(",", ".")

            # Print the currency rate
            print(f"USD to RUB currency rate: {euro_rate}")
            return float(euro_rate)
        else:
            print("Currency rate element not found.")
    else:
        print("Failed to fetch the webpage.")
        exit()


def load_invoice(bloomx_path):
    try:
        bloomx_workbook = load_workbook(bloomx_path)
        bloomx_worksheet = bloomx_workbook.active

    except Exception as e:
        print(f"Error has occured: {e}")
        exit()

    return bloomx_workbook, bloomx_worksheet


def bloomx_upd(bloomx_workbook, bloomx_worksheet, euro_rate):
    total_rows = bloomx_worksheet.max_row
    code_name = str(bloomx_worksheet['A11'].value).lower()
    last_flower = 0
    extra_cost_start = 0
    extra_cost_fin = 0
    customer = ''
    flower_sum = [0.0, 0.0, 0.0, 0.0]
    column_name = {
        'D11': 'КОЛ-ВО', 'E11': 'ТИП', 'L11': 'ЦЕНА, EUR',
        'M11': 'СУММА ЦВЕТОК, EUR', 'Q11': 'ДОП РАСХОД, EUR',
        'R11': 'ЦВЕТОК И ДОП РАСХОД, EUR', 'S11': 'КУРС EUR',
        'T11': 'СУММА ЦВЕТОК, РУБ', 'U11': 'ТРАНСПОРТ, РУБ',
        'V11': 'ИТОГО, РУБ', 'W11': 'ЦЕНА, РУБ'
    }
    sum_total = ['Q', 'R', 'T', 'U', 'V']

    for row in range(12, total_rows):
        if (
            bloomx_worksheet[f'D{row}'].value is None
            and
            last_flower == 0
        ):
            last_flower = row

            if 'Commission' in str(bloomx_worksheet[f'F{row+4}'].value):
                extra_cost_start = row + 4

            else:
                print("Commision start row wasn't found!")
                exit()

            if '9%' in str(bloomx_worksheet[f'F{row+14}'].value):
                extra_cost_fin = row + 14

            else:
                print("Commision final row wasn't found!")
                exit()

            if 'Total' in str(bloomx_worksheet[f'K{row+16}'].value):
                flower_sum[0] = float(bloomx_worksheet[f'M{row+16}'].value)

            else:
                print("Total row wasn't found!")
                exit()

            if 'Subtotal' in str(bloomx_worksheet[f'K{row+2}'].value):
                flower_sum[1] = float(bloomx_worksheet[f'M{row+2}'].value)

            else:
                print("Subtotal row wasn't found!")
                exit()

        elif (
            row >= extra_cost_start
            and
            row <= extra_cost_fin
            and
            extra_cost_start != 0
            and
            extra_cost_fin != 0
        ):
            extra = float(bloomx_worksheet[f'M{row}'].value)
            flower_sum[2] = flower_sum[2] + extra

    extra_ratio = flower_sum[2]/flower_sum[1]

    for code, info in bloomx_codes.items():
        if code in code_name:
            customer = info[0]
            euro_rate_upd = (float(euro_rate) + info[2]) * info[3]
            truck_cost = float(input('Total logistics cost:\n'))
            truck_ratio = truck_cost/flower_sum[0]

            for row in range(12, last_flower):
                subtotal = float(bloomx_worksheet[f'M{row}'].value)
                extra_subtotal = subtotal * extra_ratio
                total_eur = extra_subtotal + subtotal
                truck_local = truck_ratio * total_eur
                total_rub = euro_rate_upd * total_eur
                amount = float(bloomx_worksheet[f'D{row}'].value)
                price = (truck_local + total_rub) / amount
                bloomx_worksheet[f'Q{row}'] = round(extra_subtotal, 3)
                bloomx_worksheet[f'R{row}'] = round(subtotal + extra_subtotal, 3)
                bloomx_worksheet[f'S{row}'] = round(euro_rate_upd, 3)
                bloomx_worksheet[f'T{row}'] = round(total_rub, 3)
                bloomx_worksheet[f'U{row}'] = round(truck_local, 3)
                bloomx_worksheet[f'V{row}'] = round(truck_local + total_rub, 3)
                bloomx_worksheet[f'W{row}'] = round(price, 3)

                flower_sum[3] = flower_sum[3] + total_eur

            break

    for cell, name in column_name.items():
        bloomx_worksheet[cell] = name

    for cell in sum_total:
        bloomx_worksheet[f'{cell}{last_flower}'] = f'=SUM({cell}12:{cell}{last_flower-1})'

    if round(flower_sum[0], 2) == round(flower_sum[3], 2):
        print(f'Total flower sale in EUR: {flower_sum[3]}')
        bloomx_workbook.save(f'Bloomx {customer}.xlsx')
    else:
        print(flower_sum[0], '/', flower_sum[3])
        print("Total in EUR didn't match!")
        exit()


def start():
    invoice_name = input('Name of the invoice:\n')
    bloomx_path = f'{invoice_name}.xlsx'
    bloomx_workbook, bloomx_worksheet = load_invoice(bloomx_path)
    euro_rate = parsing_currency()
    bloomx_upd(bloomx_workbook, bloomx_worksheet, euro_rate)


warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
RGB.__set__ = __rgb_set_fixed__
start()
